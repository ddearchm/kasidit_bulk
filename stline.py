# 📦 IMPORT & CONFIG
import streamlit as st
import pandas as pd
import os, re
from io import BytesIO
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from rapidfuzz import fuzz
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName

st.set_page_config(page_title="Survey Column Builder", layout="wide")
st.title("📋 สร้างแบบสอบถาม (Excel และ PDF)")

# 🎯 SETUP SESSION STATE
if "custom_questions" not in st.session_state:
    st.session_state.custom_questions = []
if "custom_product_details" not in st.session_state:
    st.session_state.custom_product_details = []

# 🌟 FUZZY MATCH
FUZZY_MATCH_THRESHOLD = 80

def clean_question(text):
    text = str(text).strip().lower()
    return re.sub(r"\d+$", "", text)

def find_q_group(base_question, sheets_data):
    """
    คง logic เดิมไว้: หา group จากคลังคำถามของ business type ที่เลือก (sheets_data)
    """
    base = clean_question(base_question)
    best_score, best_group = 0, "N/A"
    for df in sheets_data.values():
        if "standard_question_th" in df.columns and "q_group" in df.columns:
            df = df.copy()
            df["standard_clean"] = df["standard_question_th"].astype(str).apply(clean_question)
            for _, row in df.iterrows():
                ref_q = row["standard_clean"]
                score = max(fuzz.partial_ratio(base, ref_q), fuzz.token_sort_ratio(base, ref_q))
                if score > best_score and score >= FUZZY_MATCH_THRESHOLD:
                    best_score = score
                    best_group = str(row["q_group"])
    return best_group

# 🔐 ป้องกัน duplicate column names
seen_labels = set()
def generate_unique_label(base, i, qty):
    raw = f"{base}#{i}" if qty > 1 else base
    label = raw
    count = 2
    while label in seen_labels:
        label = f"{raw}#{count}"
        count += 1
    seen_labels.add(label)
    return label

# 🧰 QUESTION BANK (ใส่คำถามจริงของคุณแทนที่ตัวอย่างด้านล่าง)
# โครงสร้าง: QUESTION_BANK[BUSINESS_TYPE][SHEET_NAME] = list ของ dict ที่มี standard_question_th, q_group
# sheet name ใช้ชื่อเดียวกับตอนอ่านจาก Excel เดิม เช่น "Respondent Profile", "Customer & Market", "Product List", "Product & Details"
QUESTION_BANK = {
    "Bulk transformer": {
        "Respondent Profile": [
            {"standard_question_th": "ชื่อ", "q_group": "Respondent Profile"},
            {"standard_question_th": "ชื่อธุรกิจ", "q_group": "Respondent Profile"},
            {"standard_question_th": "จังหวัด (ตามที่อยู่)", "q_group": "Respondent Profile"},
            {"standard_question_th": "เบอร์โทร", "q_group": "Respondent Profile"},
            {"standard_question_th": "เพศ", "q_group": "Respondent Profile"},
            {"standard_question_th": "อายุ", "q_group": "Respondent Profile"},                        
            {"standard_question_th": "ตำแหน่ง", "q_group": "Respondent Profile"},
            {"standard_question_th": "Persona", "q_group": "Respondent Profile"},

        ],
        "Customer & Market": [
            {"standard_question_th": "ประเภทงานก่อสร้างของลูกค้าหลัก", "q_group": "Customer & Market"},
            {"standard_question_th": "วิธีสื่อสารกับลูกค้าแบบออฟไลน์", "q_group": "Customer & Market"},
            {"standard_question_th": "วิธีสื่อสารกับลูกค้าแบบออนไลน์", "q_group": "Customer & Market"},
            {"standard_question_th": "ช่วงอายุของลูกค้า", "q_group": "Customer & Market"},
            {"standard_question_th": "วิธีดูแลลูกค้าประจำของร้าน", "q_group": "Customer & Market"},
            {"standard_question_th": "ระบบสะสมแต้มของตัวเอง", "q_group": "Customer & Market"},
            {"standard_question_th": "ของแจกที่ลูกค้าชอบ", "q_group": "Customer & Market"},
            {"standard_question_th": "ช่องทางการขายที่ยอดมากที่สุด", "q_group": "Customer & Market"},
            {"standard_question_th": "ช่องทางการซื้อของลูกค้าส่วนใหญ่", "q_group": "Customer & Market"},

        ],
        "Business & Strategy": [
            {"standard_question_th": "Dealer", "q_group": "Business & Strategy"},
            {"standard_question_th": "BP Model", "q_group": "Business & Strategy"},
            {"standard_question_th": "ความเป็นมาของธุรกิจ", "q_group": "Business & Strategy"},
            {"standard_question_th": "มีคนรับช่วงธุรกิจต่อหรือไม่", "q_group": "Business & Strategy"},
            {"standard_question_th": "ธุรกิจอื่นที่ทำควบคู่กัน", "q_group": "Business & Strategy"},
            {"standard_question_th": "ธุรกิจอื่นที่ทำควบคู่กัน (detail)", "q_group": "Business & Strategy"},
            {"standard_question_th": "แผนขยายธุรกิจอื่นๆ", "q_group": "Business & Strategy"},
            {"standard_question_th": "แผนขยายธุรกิจหลัก", "q_group": "Business & Strategy"},

        ],
        "Pain Points & Needs": [
            {"standard_question_th": "need ในขั้นตอนการทำงาน", "q_group": "Pain Points & Needs"},
            {"standard_question_th": "pain ในขั้นตอนการทำงาน", "q_group": "Pain Points & Needs"},
            {"standard_question_th": "แก้ไข pain ในขั้นตอนการทำงานอย่างไร", "q_group": "Pain Points & Needs"},
            {"standard_question_th": "สิ่งที่อยากให้ SCG สนับสนุน/ช่วยเหลือ", "q_group": "Pain Points & Needs"},
            {"standard_question_th": "สิ่งที่อยากให้ SCG สนับสนุน/ช่วยเหลือ (detail)", "q_group": "Pain Points & Needs"},
            
        ],
        "Product & Process": [
            {"standard_question_th": "ปูน SCG ที่ใช้ในกระบวนการผลิต", "q_group": "Product & Process"},
            {"standard_question_th": "ปัจจัยสำคัญในการเลือกซื้อปูน เสือ/SCG", "q_group": "Product & Process"},
            {"standard_question_th": "แบรนด์ขายดี", "q_group": "Product & Process"},
            {"standard_question_th": "สินค้าอื่นที่ผลิตขาย", "q_group": "Product & Process"},
            {"standard_question_th": "กลยุทธ์รักษาฐานลูกค้า และสู้กับคู่แข่ง", "q_group": "Product & Process"},
            {"standard_question_th": "ปัจจัยเสี่ยงต่อธุรกิจ", "q_group": "Product & Process"},
            {"standard_question_th": "ปัจจัยสำคัญของการผลิต", "q_group": "Product & Process"},
            {"standard_question_th": "ความสำคัญของคุณภาพปูนในกระบวนการผลิต", "q_group": "Product & Process"},
            {"standard_question_th": "คุณสมบัติปูนที่สำคัญ", "q_group": "Product & Process"},
            {"standard_question_th": "คุณสมบัติสำคัญในกระบวนการผลิตสินค้าคอนกรีตขายดี", "q_group": "Product & Process"},
            {"standard_question_th": "คุณสมบัติของสินค้าคอนกรีตขายดี", "q_group": "Product & Process"},
            {"standard_question_th": "วิธีการทำงานในส่วน Pre-Stressed", "q_group": "Product & Process"},
            {"standard_question_th": "วิธีการทำงานในส่วน RMC", "q_group": "Product & Process"},
            {"standard_question_th": "วิธีการทำงานในส่วน Non-Prestressed", "q_group": "Product & Process"},
            {"standard_question_th": "กระบวนการทำงานในโรงหล่อที่สำคัญ", "q_group": "Product & Process"},
            {"standard_question_th": "แหล่งวัตถุดิบ", "q_group": "Product & Process"},
            {"standard_question_th": "วิธีเช็คคุณภาพวัตถุดิบ", "q_group": "Product & Process"},
            {"standard_question_th": "การตรวจสอบคุณภาพสินค้า", "q_group": "Product & Process"},

        ],        
    },
    "Bag transformer": {
        "Respondent Profile": [
            {"standard_question_th": "ชื่อ", "q_group": "Respondent Profile"},
            {"standard_question_th": "ชื่อธุรกิจ", "q_group": "Respondent Profile"},
            {"standard_question_th": "จังหวัด (ตามที่อยู่)", "q_group": "Respondent Profile"},
            {"standard_question_th": "เบอร์โทร", "q_group": "Respondent Profile"},
            {"standard_question_th": "เพศ", "q_group": "Respondent Profile"},
            {"standard_question_th": "อายุ", "q_group": "Respondent Profile"},                        
            {"standard_question_th": "ตำแหน่ง", "q_group": "Respondent Profile"},
            {"standard_question_th": "Persona", "q_group": "Respondent Profile"},

        ],
        "Customer & Market": [
            {"standard_question_th": "ประเภทงานก่อสร้างของลูกค้าหลัก", "q_group": "Customer & Market"},
            {"standard_question_th": "วิธีสื่อสารกับลูกค้าแบบออฟไลน์", "q_group": "Customer & Market"},
            {"standard_question_th": "วิธีสื่อสารกับลูกค้าแบบออนไลน์", "q_group": "Customer & Market"},
        ],
        "Business & Strategy": [
            {"standard_question_th": "Dealer", "q_group": "Business & Strategy"},
            {"standard_question_th": "BP Model", "q_group": "Business & Strategy"},
            {"standard_question_th": "ความเป็นมาของธุรกิจ", "q_group": "Business & Strategy"},
            {"standard_question_th": "มีคนรับช่วงธุรกิจต่อหรือไม่", "q_group": "Business & Strategy"},
            {"standard_question_th": "ธุรกิจอื่นที่ทำควบคู่กัน", "q_group": "Business & Strategy"},
            {"standard_question_th": "แผนขยายธุรกิจอื่นๆ", "q_group": "Business & Strategy"},
            {"standard_question_th": "แผนขยายธุรกิจหลัก", "q_group": "Business & Strategy"},
        ],
        "Pain Points & Needs": [
            {"standard_question_th": "need ในขั้นตอนการทำงาน", "q_group": "Pain Points & Needs"},
            {"standard_question_th": "pain ในขั้นตอนการทำงาน", "q_group": "Pain Points & Needs"},
            {"standard_question_th": "แก้ไข pain ในขั้นตอนการทำงานอย่างไร", "q_group": "Pain Points & Needs"},
            {"standard_question_th": "สิ่งที่อยากให้ SCG สนับสนุน/ช่วยเหลือ", "q_group": "Pain Points & Needs"},
            {"standard_question_th": "สิ่งที่อยากให้ SCG สนับสนุน/ช่วยเหลือ (detail)", "q_group": "Pain Points & Needs"},
            
        ],
        "Product & Process": [
            {"standard_question_th": "ปูน SCG ที่ใช้ในกระบวนการผลิต", "q_group": "Product & Process"},
            {"standard_question_th": "ปัจจัยสำคัญในการเลือกซื้อปูน เสือ/SCG", "q_group": "Product & Process"},
            {"standard_question_th": "แบรนด์ขายดี", "q_group": "Product & Process"},
            {"standard_question_th": "สินค้าอื่นที่ผลิตขาย", "q_group": "Product & Process"},
            {"standard_question_th": "กลยุทธ์รักษาฐานลูกค้า และสู้กับคู่แข่ง", "q_group": "Product & Process"},
            {"standard_question_th": "ปัจจัยเสี่ยงต่อธุรกิจ", "q_group": "Product & Process"},
            {"standard_question_th": "ปัจจัยสำคัญของการผลิต", "q_group": "Product & Process"},
            {"standard_question_th": "ความสำคัญของคุณภาพปูนในกระบวนการผลิต", "q_group": "Product & Process"},
            {"standard_question_th": "คุณสมบัติปูนที่สำคัญ", "q_group": "Product & Process"},
            {"standard_question_th": "คุณสมบัติสำคัญในกระบวนการผลิตสินค้าคอนกรีตขายดี", "q_group": "Product & Process"},
            {"standard_question_th": "คุณสมบัติของสินค้าคอนกรีตขายดี", "q_group": "Product & Process"},
            {"standard_question_th": "วิธีการทำงานในส่วน Prestressed", "q_group": "Product & Process"},
            {"standard_question_th": "วิธีการทำงานในส่วน RMC", "q_group": "Product & Process"},
            {"standard_question_th": "วิธีการทำงานในส่วน Non-Prestressed", "q_group": "Product & Process"},
            {"standard_question_th": "กระบวนการทำงานในโรงหล่อที่สำคัญ", "q_group": "Product & Process"},
            {"standard_question_th": "แหล่งวัตถุดิบ", "q_group": "Product & Process"},
            {"standard_question_th": "วิธีเช็คคุณภาพวัตถุดิบ", "q_group": "Product & Process"},
            {"standard_question_th": "การตรวจสอบคุณภาพสินค้า", "q_group": "Product & Process"},
        ],        
    },
    "Subdealer & Bag transformer": {
        "Respondent Profile": [
            {"standard_question_th": "บริษัทรับเหมาก่อสร้าง", "q_group": "Respondent Profile"},
            {"standard_question_th": "ชื่อ", "q_group": "Respondent Profile"},
            {"standard_question_th": "เบอร์โทร", "q_group": "Respondent Profile"},
            {"standard_question_th": "เพศ", "q_group": "Respondent Profile"},
            {"standard_question_th": "อายุ", "q_group": "Respondent Profile"},
            {"standard_question_th": "ตำแหน่ง", "q_group": "Respondent Profile"},
            {"standard_question_th": "จังหวัด (ตามที่อยู่)", "q_group": "Respondent Profile"},
        ],
        "Customer & Market": [
            {"standard_question_th": "ประเภทงานก่อสร้างของลูกค้าหลัก", "q_group": "Customer & Market"},
            {"standard_question_th": "ยอดซื้อเฉลี่ยของลูกค้า (บาท/บิล)", "q_group": "Customer & Market"},
            {"standard_question_th": "สัดส่วนลูกค้าโทรสั่ง", "q_group": "Customer & Market"},
            {"standard_question_th": "สัดส่วนลูกค้าไลน์สั่ง", "q_group": "Customer & Market"},
            {"standard_question_th": "สัดส่วนลูกค้าสั่งที่ร้าน", "q_group": "Customer & Market"},
            {"standard_question_th": "กลุ่มลูกค้าประจำของร้าน", "q_group": "Customer & Market"},
            {"standard_question_th": "วิธีดูแลลูกค้าประจำของร้าน", "q_group": "Customer & Market"},
        ],
        "Business & Strategy": [
            {"standard_question_th": "Dealer", "q_group": "Business & Strategy"},
            {"standard_question_th": "สถานการณ์ตลาด", "q_group": "Business & Strategy"},
            {"standard_question_th": "สถานการณ์แข่งขันด้านราคา", "q_group": "Business & Strategy"},
            {"standard_question_th": "Price Gap ที่เหมาะสม", "q_group": "Business & Strategy"},
            {"standard_question_th": "สรุปช่องทางที่ลูกค้าสั่งซื้อสินค้า", "q_group": "Business & Strategy"},
            {"standard_question_th": "รูปแบบการชำระเงิน", "q_group": "Business & Strategy"},
            {"standard_question_th": "แฟนพันธ์แท้ปูนเสือ/SCG", "q_group": "Business & Strategy"},
            {"standard_question_th": "Capacity หน้าร้าน (ตัน)", "q_group": "Business & Strategy"},
            {"standard_question_th": "Capacity รวมทั้งร้าน (ตัน)", "q_group": "Business & Strategy"},
            {"standard_question_th": "มีคนรับช่วงธุรกิจต่อหรือไม่", "q_group": "Business & Strategy"},
            {"standard_question_th": "ธุรกิจอื่นที่ทำควบคู่กัน", "q_group": "Business & Strategy"},
            {"standard_question_th": "แผนขยายธุรกิจ", "q_group": "Business & Strategy"},
            {"standard_question_th": "ทำธุรกิจโรงหล่อควบคู่ร้านวัสดุก่อสร้าง", "q_group": "Business & Strategy"},            
        ],
        "Pain Points & Needs": [
            {"standard_question_th": "need ในขั้นตอนการทำงาน", "q_group": "Pain Points & Needs"},
            {"standard_question_th": "pain ในขั้นตอนการทำงาน", "q_group": "Pain Points & Needs"},
            {"standard_question_th": "แก้ไข pain ในขั้นตอนการทำงานอย่างไร", "q_group": "Pain Points & Needs"},
            {"standard_question_th": "สิ่งที่อยากให้ SCG สนับสนุน/ช่วยเหลือ", "q_group": "Pain Points & Needs"},
            {"standard_question_th": "สิ่งที่อยากให้ SCG สนับสนุน/ช่วยเหลือ (detail)", "q_group": "Pain Points & Needs"},
            
        ],
        "Product & Process": [
            {"standard_question_th": "ปูน SCG ที่ใช้ในกระบวนการผลิต", "q_group": "Business & Strategy"},
            {"standard_question_th": "สินค้าหลักที่ผลิต", "q_group": "Business & Strategy"},
            {"standard_question_th": "สินค้าขายดี", "q_group": "Business & Strategy"},
            {"standard_question_th": "กลุ่มลูกค้าหลักของธุรกิจโรงหล่อ", "q_group": "Business & Strategy"},
            {"standard_question_th": "แหล่งวัตถุดิบและวิธีเช็คคุณภาพ", "q_group": "Business & Strategy"},
            {"standard_question_th": "ระบบขายหน้าร้าน", "q_group": "Product & Process"},
            {"standard_question_th": "แบรนด์ที่ขายดี", "q_group": "Product & Process"},
            {"standard_question_th": "ปัจจัยสำคัญในการเลือกซื้อปูน เสือ/SCG", "q_group": "Product & Process"},
            {"standard_question_th": "กลยุทธ์รักษาฐานลูกค้า และสู้กับคู่แข่ง", "q_group": "Product & Process"},
            {"standard_question_th": "ปัจจัยสำคัญของการผลิต", "q_group": "Product & Process"},            
        ],   
        "Product List": [
            {"standard_question_th": "ก่อ-Grey", "q_group": "Product & Details"},
            {"standard_question_th": "ก่อ-Mortar", "q_group": "Product & Details"},
            {"standard_question_th": "ก่อ-Mortar-LW", "q_group": "Product & Details"},
            {"standard_question_th": "ฉาบ-Grey", "q_group": "Product & Details"},
            {"standard_question_th": "ฉาบ-Mortar", "q_group": "Product & Details"},
            {"standard_question_th": "ฉาบ-Mortar-LW", "q_group": "Product & Details"},
            {"standard_question_th": "ฉาบ-Grey-จับเซี๊ยม", "q_group": "Product & Details"},
            {"standard_question_th": "ฉาบ-Mortar-จับเซี๊ยม", "q_group": "Product & Details"},
            {"standard_question_th": "ฉาบบาง-Mortar-สกิมโค้ท", "q_group": "Product & Details"},
            {"standard_question_th": "เทโครงสร้าง-Grey", "q_group": "Product & Details"},
            {"standard_question_th": "เทโครงสร้าง-Mortar", "q_group": "Product & Details"},
            {"standard_question_th": "เทโครงสร้าง-RMC", "q_group": "Product & Details"},
            {"standard_question_th": "เทเสาเอ็น-Grey", "q_group": "Product & Details"},
            {"standard_question_th": "เทเสาเอ็น-Mortar", "q_group": "Product & Details"},
            {"standard_question_th": "เทปรับพื้น-Grey", "q_group": "Product & Details"},
            {"standard_question_th": "เทปรับพื้น-Mortar", "q_group": "Product & Details"},
            {"standard_question_th": "เทปรับพื้น-RMC", "q_group": "Product & Details"},
            {"standard_question_th": "ปูกระเบื้อง-Mortar-TA", "q_group": "Product & Details"},
            {"standard_question_th": "ปูกระเบื้อง-Mortar-TG", "q_group": "Product & Details"},
            {"standard_question_th": "ผนัง-อิฐมอญ", "q_group": "Product & Details"},
            {"standard_question_th": "ผนัง-อิฐบล็อก", "q_group": "Product & Details"},
            {"standard_question_th": "ผนัง-อิฐมวลเบา", "q_group": "Product & Details"},
            {"standard_question_th": "ผนัง-CLC", "q_group": "Product & Details"},
            {"standard_question_th": "ผนัง-Wall system", "q_group": "Product & Details"},
            {"standard_question_th": "สี-รองพื้น", "q_group": "Product & Details"},
            {"standard_question_th": "สี-สีจริง", "q_group": "Product & Details"},
            {"standard_question_th": "อื่นๆ-Water proof", "q_group": "Product & Details"},
            {"standard_question_th": "อื่นๆ-Non shrink", "q_group": "Product & Details"},
            {"standard_question_th": "อื่นๆ-White", "q_group": "Product & Details"},
        ],
        "Product & Details": [
            {"standard_question_th": "ยี่ห้อ", "q_group": "Product & Details"},
            {"standard_question_th": "ราคาหน้าร้าน", "q_group": "Product & Details"},
            {"standard_question_th": "ราคาทุน", "q_group": "Product & Details"},
            {"standard_question_th": "สต็อก", "q_group": "Product & Details"},
        ],
        "Special Topic": [
            {"standard_question_th": "Giant Banner", "q_group": "Special Topic"},
            {"standard_question_th": "รูปแบบบิลที่ใช้ในแต้มปูน", "q_group": "Special Topic"},
            {"standard_question_th": "เหตุผลที่เป็นแฟนพันธุ์แท้ปูนเสือ/SCG", "q_group": "Special Topic"},
        ],
    },
    "Contractor": {
         "Respondent Profile": [
            {"standard_question_th": "ชื่อ", "q_group": "Respondent Profile"},
            {"standard_question_th": "ชื่อเล่น", "q_group": "Respondent Profile"},
            {"standard_question_th": "เพศ", "q_group": "Respondent Profile"},
            {"standard_question_th": "อายุ", "q_group": "Respondent Profile"},
            {"standard_question_th": "ตำแหน่ง", "q_group": "Respondent Profile"},
            {"standard_question_th": "จำนวนทีมงาน", "q_group": "Respondent Profile"},
            {"standard_question_th": "ประสบการณ์ทำงาน", "q_group": "Respondent Profile"},
            {"standard_question_th": "กิจวัตรประจำวันและงานอดิเรก", "q_group": "Respondent Profile"},
            {"standard_question_th": "เป้าหมายชีวิต", "q_group": "Respondent Profile"},
            {"standard_question_th": "สิ่งที่อยากพัฒนาเพื่อให้ธุรกิจดีขึ้น", "q_group": "Respondent Profile"},
            {"standard_question_th": "ประวัติการศึกษาและจุดเริ่มต้นการทำงาน", "q_group": "Respondent Profile"},
            {"standard_question_th": "สื่อที่ใช้ในการรับข้อมูล", "q_group": "Respondent Profile"},
            {"standard_question_th": "lifestyle", "q_group": "Respondent Profile"},
        ],
        "Customer's Journey": [
            {"standard_question_th": "ค้นหาข้อมูลก่อนซื้ออย่างไร", "q_group": "Customer's Journey"},
            {"standard_question_th": "ปัจจัยสำคัญในการเลือกซื้อปูน", "q_group": "Customer's Journey"},
            {"standard_question_th": "ร้านค้าที่ส่งผลต่อการซื้อปูน", "q_group": "Customer's Journey"},
            {"standard_question_th": "สื่อ/ช่องทางที่ส่งผลต่อการซื้อปูน", "q_group": "Customer's Journey"},
            {"standard_question_th": "ร้าน Modern Trade ที่ซื้อวัสดุุ", "q_group": "Customer's Journey"},
            {"standard_question_th": "สินค้าที่ซื้อมากที่สุดจาก Modern Trade", "q_group": "Customer's Journey"},
            {"standard_question_th": "โมเดิร์นเทรดที่เป็นสมาชิกในการสะสมแต้ม", "q_group": "Customer's Journey"},
            {"standard_question_th": "ระบบสะสมแต้มนี้ตอบโจทย์คุณในด้านใด", "q_group": "Customer's Journey"},
            {"standard_question_th": "สิ่งที่ไม่ประทับใจ", "q_group": "Customer's Journey"},
            {"standard_question_th": "รู้สึกว่าการสะสมคะแนนยุ่งยากหรือไม่", "q_group": "Customer's Journey"},
            {"standard_question_th": "วิธีสะสมแต้ม", "q_group": "Customer's Journey"},
            {"standard_question_th": "รูปแบบบิลที่ใช้ในแต้มปูน", "q_group": "Customer's Journey"},
            {"standard_question_th": "เหตุผลที่เป็นแฟนพันธุ์แท้ปูนเสือ/SCG", "q_group": "Customer's Journey"},
        ],
        "Customer & Market": [
            {"standard_question_th": "ใครเป็นผู้ตัดสินใจซื้อวัสดุก่อสร้าง", "q_group": "Customer & Market"},
            {"standard_question_th": "ประเภทงานก่อสร้างที่ให้บริการเป็นหลัก", "q_group": "Customer & Market"},
            {"standard_question_th": "แบรนด์ใดที่คุณมองว่าใกล้เคียงกับปูนเสือ/SCG", "q_group": "Customer & Market"},
            {"standard_question_th": "วิธีการสั่งซื้อปูนและวัสดุก่อสร้าง", "q_group": "Customer & Market"},
            {"standard_question_th": "%สั่งซื้อปูนและวัสดุก่อสร้างทางโทรศัพท์", "q_group": "Customer & Market"},
            {"standard_question_th": "%สั่งซื้อปูนและวัสดุก่อสร้างทางไลน์", "q_group": "Customer & Market"},
            {"standard_question_th": "%สั่งซื้อปูนและวัสดุก่อสร้างที่หน้าร้าน", "q_group": "Customer & Market"},
            {"standard_question_th": "วิธีการจ่ายเงิน (เงินสด, เครดิต, เงินสดและเครดิต)", "q_group": "Customer & Market"},
            {"standard_question_th": "ยอดซื้อปูน และวัสดุก่อสร้างโดยเฉลี่ย (บาทต่อบิล)", "q_group": "Customer & Market"},            
            {"standard_question_th": "สิ่งที่ SCG มีแต่เจ้าอื่นไม่มี", "q_group": "Customer & Market"},
        ],
        "Business & Strategy": [
            {"standard_question_th": "ชื่อบริษัทรับเหมาก่อสร้าง", "q_group": "Business & Strategy"},
            {"standard_question_th": "ชื่อโครงการในวันที่เข้าสัมภาษณ์", "q_group": "Business & Strategy"},
            {"standard_question_th": "ภาค (ตามที่อยู่)", "q_group": "Business & Strategy"},
            {"standard_question_th": "จังหวัด (ตามที่อยู่)", "q_group": "Business & Strategy"},
            {"standard_question_th": "จังหวัดที่รับบริการ", "q_group": "Business & Strategy"},
            {"standard_question_th": "ร้านประจำที่ซื้อวัสดุก่อสร้าง", "q_group": "Business & Strategy"},            
            {"standard_question_th": "วิธีคิดค่าบริการงานก่อสร้าง", "q_group": "Business & Strategy"},
            {"standard_question_th": "มูลค่างานต่อปี", "q_group": "Business & Strategy"},
            {"standard_question_th": "ร้านประจำที่ซื้อปูน เสือ/SCG", "q_group": "Business & Strategy"},            
            {"standard_question_th": "ความถี่ในการสั่งปูน (ครั้ง/สัปดาห์)", "q_group": "Business & Strategy"},
            {"standard_question_th": "ความถี่ในการเข้าร้านวัสดุก่อสร้าง (ครั้ง/สัปดาห์)", "q_group": "Business & Strategy"},
            {"standard_question_th": "มีคนรับช่วงธุรกิจต่อหรือไม่", "q_group": "Business & Strategy"},
            {"standard_question_th": "ธุรกิจอื่นที่ทำควบคู่กัน", "q_group": "Business & Strategy"},
        ],
        "Pain Points & Needs": [
            {"standard_question_th": "ปัญหาที่พบในการทำงานของช่าง", "q_group": "Pain Points & Needs"},
            {"standard_question_th": "ปัญหาที่พบในการจัดซื้อวัสดุก่อสร้าง", "q_group": "Pain Points & Needs"},
            {"standard_question_th": "สิ่งที่อยากให้ SCG พัฒนา", "q_group": "Pain Points & Needs"},
            {"standard_question_th": "สิ่งที่อยากให้ SCG สนับสนุน/ช่วยเหลือ", "q_group": "Pain Points & Needs"},                        
        ],
        "Product & Process": [
            {"standard_question_th": "ปูนเสือ/SCG ที่ซื้อไป นิยมใช้กับงานประเภทใด", "q_group": "Product & Process"},
            {"standard_question_th": "สินค้าที่มักซื้อคู่กับปูน", "q_group": "Product & Process"},
            {"standard_question_th": "ปริมาณสินค้าที่มักซื้อคู่กับปูน", "q_group": "Product & Process"},
            {"standard_question_th": "สนใจทดลองใช้สินค้า scg หรือไม่", "q_group": "Product & Process"},
            {"standard_question_th": "ปูน หรือสินค้า scg ที่สนใจทดลองใช้", "q_group": "Product & Process"},            
        ],
        # Contractor อาจไม่มี cross-product ก็ได้ — ถ้าไม่มี ก็ลบสองชีตนี้ออก
        "Product List": [
            {"standard_question_th": "ก่อ-Grey", "q_group": "Product & Details"},
            {"standard_question_th": "ก่อ-Mortar", "q_group": "Product & Details"},
            {"standard_question_th": "ก่อ-Mortar-LW", "q_group": "Product & Details"},
            {"standard_question_th": "ฉาบ-Grey", "q_group": "Product & Details"},
            {"standard_question_th": "ฉาบ-Mortar", "q_group": "Product & Details"},
            {"standard_question_th": "ฉาบ-Mortar-LW", "q_group": "Product & Details"},
            {"standard_question_th": "ฉาบ-Grey-จับเซี๊ยม", "q_group": "Product & Details"},
            {"standard_question_th": "ฉาบ-Mortar-จับเซี๊ยม", "q_group": "Product & Details"},
            {"standard_question_th": "ฉาบบาง-Mortar-สกิมโค้ท", "q_group": "Product & Details"},
            {"standard_question_th": "เทโครงสร้าง-Grey", "q_group": "Product & Details"},
            {"standard_question_th": "เทโครงสร้าง-Mortar", "q_group": "Product & Details"},
            {"standard_question_th": "เทโครงสร้าง-RMC", "q_group": "Product & Details"},
            {"standard_question_th": "เทเสาเอ็น-Grey", "q_group": "Product & Details"},
            {"standard_question_th": "เทเสาเอ็น-Mortar", "q_group": "Product & Details"},
            {"standard_question_th": "เทปรับพื้น-Grey", "q_group": "Product & Details"},
            {"standard_question_th": "เทปรับพื้น-Mortar", "q_group": "Product & Details"},
            {"standard_question_th": "เทปรับพื้น-RMC", "q_group": "Product & Details"},
            {"standard_question_th": "ปูกระเบื้อง-Mortar-TA", "q_group": "Product & Details"},
            {"standard_question_th": "ปูกระเบื้อง-Mortar-TG", "q_group": "Product & Details"},
            {"standard_question_th": "ผนัง-อิฐมอญ", "q_group": "Product & Details"},
            {"standard_question_th": "ผนัง-อิฐบล็อก", "q_group": "Product & Details"},
            {"standard_question_th": "ผนัง-อิฐมวลเบา", "q_group": "Product & Details"},
            {"standard_question_th": "ผนัง-CLC", "q_group": "Product & Details"},
            {"standard_question_th": "ผนัง-Wall system", "q_group": "Product & Details"},
            {"standard_question_th": "สี-รองพื้น", "q_group": "Product & Details"},
            {"standard_question_th": "สี-สีจริง", "q_group": "Product & Details"},
            {"standard_question_th": "อื่นๆ-Water proof", "q_group": "Product & Details"},
            {"standard_question_th": "อื่นๆ-Non shrink", "q_group": "Product & Details"},
            {"standard_question_th": "อื่นๆ-White", "q_group": "Product & Details"},
        ],
        "Product & Details": [
            {"standard_question_th": "ยี่ห้อ/รุ่น", "q_group": "Product & Details"},
            {"standard_question_th": "ใช้แตกต่างกันอย่างไร?", "q_group": "Product & Details"},
            {"standard_question_th": "ใคร Spec/ ใครเลือก", "q_group": "Product & Details"},
            {"standard_question_th": "ร้านที่ซื้อ", "q_group": "Product & Details"},
            {"standard_question_th": "ปัจจัยการเลือกร้าน", "q_group": "Product & Details"},
            {"standard_question_th": "ปัจจัยเลือกแบรนด์", "q_group": "Product & Details"},
            {"standard_question_th": "ราคา (บาท/ถุง)", "q_group": "Product & Details"},
            {"standard_question_th": "ปริมาณการซื้อต่อครั้ง", "q_group": "Product & Details"},
        ],
    },
}

BUSINESS_TYPES = list(QUESTION_BANK.keys())
st.markdown("""<style>.heading-lg{ font-size:1.25rem; font-weight:700; margin:8px 0 4px; }</style>""", unsafe_allow_html=True)
# 🧭 เลือก Business Type ก่อน (แทนที่การอัปโหลดไฟล์)
# หัวข้อใหญ่ (จะใหญ่กว่า markdown ปกติ)
st.subheader("🏷️ เลือก BUSINESS TYPE ก่อนเริ่ม")

try:
    # เวอร์ชันใหม่ของ Streamlit
    biz = st.selectbox(
        "",
        options=list(QUESTION_BANK.keys()),
        index=None,
        placeholder="— เลือก BUSINESS_TYPE —",
        label_visibility="collapsed",
    )
except TypeError:
    # เวอร์ชันเก่า: ทำ placeholder เอง
    PLACEHOLDER = "— เลือก BUSINESS_TYPE —"
    biz = st.selectbox(
        "",
        options=[PLACEHOLDER] + list(QUESTION_BANK.keys()),
        index=0,
        label_visibility="collapsed",
    )
    if biz == PLACEHOLDER:
        st.info("👆 กรุณาเลือก BUSINESS TYPE เพื่อสร้างคำถาม")
        st.stop()

# เวอร์ชันใหม่: ถ้ายังไม่เลือก จะเป็น None
if not biz:
    st.info("👆 กรุณาเลือก BUSINESS TYPE เพื่อสร้างคำถาม")
    st.stop()


# 📚 แปลง QUESTION_BANK -> sheets_data (โครงเดียวกับไฟล์ Excel เดิม)
PRODUCT_SHEETS = {"Product List", "Product & Details"}

def build_sheets_data_from_bank(bank_for_biz: dict) -> dict:
    sheets = {}
    for sheet_name, rows in (bank_for_biz or {}).items():
        # ข้ามชีตที่ว่างเปล่า
        if rows is None or (isinstance(rows, list) and len(rows) == 0):
            continue

        df = pd.DataFrame(rows)

        # ถ้าเป็น list[str] จะได้คอลัมน์ชื่อ 0 มา -> แปลงให้ถูก
        if "standard_question_th" not in df.columns:
            if 0 in df.columns:  # เคส list[str]
                df = pd.DataFrame({"standard_question_th": df[0].astype(str).str.strip()})
            else:
                # ไม่มีคำถามเลยก็ข้ามไป
                if df.empty:
                    continue
                raise ValueError(f"{sheet_name}: missing 'standard_question_th'")

        if "q_group" not in df.columns:
            df["q_group"] = "Product & Details" if sheet_name in PRODUCT_SHEETS else sheet_name

        df = df[["standard_question_th", "q_group"]].copy()
        df["standard_question_th"] = df["standard_question_th"].astype(str).str.strip()
        df["q_group"] = df["q_group"].astype(str).str.strip()

        # ข้ามแถวที่คำถามว่าง
        df = df[df["standard_question_th"] != ""]
        if not df.empty:
            sheets[sheet_name] = df

    return sheets


sheets_data = build_sheets_data_from_bank(QUESTION_BANK.get(biz, {}))


# ตรวจว่ามี cross-product ไหม
is_cross = "Product List" in sheets_data and "Product & Details" in sheets_data

# =========================
#   UI เลือกคำถาม (มาตรฐานก่อน → ค่อย Product)
# =========================

st.subheader("📌 คำถามที่ต้องการในการเก็บข้อมูล")
selected_questions = []
# ลำดับกลุ่มมาตรฐาน (สำหรับหน้าจอเลือกคำถาม)
ORDER_STANDARD_GROUPS = [
    "Respondent Profile",
    "Customer's Journey",
    "Customer & Market",
    "Business & Strategy",
    "Pain Points & Needs",
    "Product & Process",
    "Special Topic",
]

# วนตามลำดับที่กำหนดไว้
for sheet_name in ORDER_STANDARD_GROUPS:
    if sheet_name in sheets_data and "standard_question_th" in sheets_data[sheet_name].columns:
        df = sheets_data[sheet_name]
        st.markdown(f"<h4 style='margin:6px 0;text-decoration:underline;'>📑 {sheet_name}</h4>", unsafe_allow_html=True)
        for i, row in df.iterrows():
            q = str(row["standard_question_th"])
            if pd.notna(q) and q.strip():
                if st.checkbox(q, key=f"{sheet_name}_{i}"):
                    qty = st.number_input(
                        f"🔢 จำนวน: {q[:30]}",
                        1, 20, 1, 1,
                        key=f"{sheet_name}_{i}_qty"
                    )
                    # group จากแหล่งข้อมูล ถ้าไม่มีให้เป็น N/A (ยังมี fuzzy สำรองตอน export)
                    selected_questions.append({
                        "Question": q.strip(),
                        "Quantity": qty,
                        "Group": row.get("q_group", "N/A")
                    })

# —— หลังจากนั้นค่อย “กลุ่ม Product สำหรับ cross” ——
selected_products, selected_details = [], []
if is_cross:
    st.subheader("📑 กลุ่ม Product List")

    # Product List มาก่อน
    st.markdown("<div class='heading-lg' style='text-decoration: underline;'>📦 Product List</div>", unsafe_allow_html=True)
    prod_df = sheets_data["Product List"]

    # ให้ 2 กลุ่มนี้ติ๊กทั้งหมดเป็นค่าเริ่มต้น
    DEFAULT_SELECT_ALL_BIZ = {"Subdealer & Bag transformer", "Contractor"}
    default_select_all = (biz in DEFAULT_SELECT_ALL_BIZ)

    # ทำ prefix ให้ key ไม่ชนกันข้าม business type
    prod_prefix = f"prod_{biz.replace(' ', '_')}"

    # init ครั้งแรกของ Product List (ต่อ business type)
    if st.session_state.get(f"{prod_prefix}_initialized") is None:
        st.session_state[f"{prod_prefix}_select_all"] = default_select_all
        st.session_state[f"{prod_prefix}_select_all_prev"] = default_select_all
        # ตั้งค่า checkbox รายการสินค้าให้ตรงกับ select_all ตอนเริ่ม
        for i in range(len(prod_df)):
            st.session_state[f"{prod_prefix}_{i}"] = default_select_all
        st.session_state[f"{prod_prefix}_initialized"] = True

    # ปุ่ม Select All
    st.checkbox("✅ เลือกทั้งหมด", key=f"{prod_prefix}_select_all")

    # ถ้า select_all เปลี่ยนค่า → sync ทุกกล่อง
    if st.session_state[f"{prod_prefix}_select_all_prev"] != st.session_state[f"{prod_prefix}_select_all"]:
        new_val = st.session_state[f"{prod_prefix}_select_all"]
        for i in range(len(prod_df)):
            st.session_state[f"{prod_prefix}_{i}"] = new_val
        st.session_state[f"{prod_prefix}_select_all_prev"] = new_val

    # วาดรายการสินค้า
    for i, row in prod_df.iterrows():
        q = str(row["standard_question_th"]).strip()
        if not q:
            continue
        checked = st.checkbox(q, key=f"{prod_prefix}_{i}")
        if checked:
            qty = st.number_input(
                f"🔢 จำนวน: {q}",
                min_value=1, max_value=20, value=1, step=1,
                key=f"{prod_prefix}_qty_{i}"
            )
            selected_products.append({"name": q, "qty": qty})
    

    # แล้วค่อย Product & Details
    st.markdown("<div class='heading-lg' style='text-decoration: underline;'>🧾 Product & Details</div>", unsafe_allow_html=True)
    for i, row in sheets_data["Product & Details"].iterrows():
        q = str(row["standard_question_th"])
        if pd.notna(q) and q.strip():
            if st.checkbox(q, key=f"detail_{i}"):
                selected_details.append(q.strip())

    with st.expander("➕ เพิ่มคำถามเกี่ยวกับสินค้า (Product Details)"):
        custom_detail = st.text_input("กรอกคำถามเกี่ยวกับสินค้า", key="custom_detail_input")
        if st.button("➕ เพิ่มคำถามเกี่ยวกับสินค้า"):
            if custom_detail.strip():
                st.session_state.custom_product_details.append(custom_detail.strip())
                st.success(f"✅ เพิ่มคำถามสินค้า \"{custom_detail.strip()}\" แล้วเรียบร้อย")
                st.info("หากต้องการเพิ่มคำถามอื่นๆ สามารถกรอกและกด 'เพิ่มคำถามเกี่ยวกับสินค้า' ได้เลย")
            else:
                st.warning("กรุณากรอกคำถาม")

# เติม custom product details เข้าไป
selected_details += st.session_state.custom_product_details

# ✍️ Custom Questions (ยังอยู่หลังกลุ่มมาตรฐาน)
st.subheader("✍️ เพิ่มคำถามเอง ")
with st.expander("✍️ เพิ่มคำถามเอง กดที่นี่"):
    custom_q = st.text_input("กรอกคำถามที่ต้องการเพิ่ม", key="custom_question_input")
    custom_q_qty = st.number_input("จำนวน Column ที่ต้องการ", 1, 20, 1, 1, key="custom_question_qty")
    custom_q_group = st.selectbox(
        "เพิ่มคำถามนี้ในกลุ่มใด? (q_group)",
        options=[            
            "Respondent Profile",
            "Customer & Market",
            "Customer's Journey",
            "Business & Strategy",
            "Pain Points & Needs",
            "Product & Process",
            "Product & Details",
            "Special Topic"            
        ],
        index=1,
        key="custom_question_group"
    )
    if st.button("➕ เพิ่มคำถามนี้"):
        if custom_q.strip():
            st.session_state.custom_questions.append({
                "Question": custom_q.strip(),
                "Quantity": custom_q_qty,
                "Group": custom_q_group
            })
            st.success(f"✅ เพิ่มคำถาม \"{custom_q.strip()}\" เข้า group \"{custom_q_group}\" แล้ว!")
            st.info("หากต้องการเพิ่มคำถามอื่นๆ สามารถกรอกและกด 'เพิ่มคำถามนี้' ได้เลย")
        else:
            st.warning("กรุณากรอกคำถาม")

# รวม custom เข้าไปด้วย
for item in st.session_state.custom_questions:
    selected_questions.append({
        "Question": item["Question"],
        "Quantity": item["Quantity"],
        "Group": item.get("Group", "N/A")
    })


# =========================
#   GENERATE EXPORT
# =========================
if st.button("📅 สร้างและดาวน์โหลด Excel + PDF"):
    columns, qgroup_row, question_row, pdf_rows = [], [], [], []
    seen_labels.clear()

    # ✅ Group questions (ยังคง logic เดิม + fuzzy สำรองจากคลังเดียวกัน)
    grouped_questions_by_group = {}
    unmatched_questions = []

    for q in selected_questions:
        base_q = q["Question"]
        # ถ้า group ใส่มาแล้ว ใช้เลย; ถ้าไม่ ก็ใช้ find_q_group จาก sheets_data
        group = q.get("Group") if q.get("Group") not in [None, "", "N/A"] else find_q_group(base_q, sheets_data)
        item = {"question": base_q, "qty": q["Quantity"], "group": group}
        if group == "N/A":
            unmatched_questions.append(item)
        else:
            grouped_questions_by_group.setdefault(group, []).append(item)

    # ✅ ลำดับ group ที่ต้องการ
    preferred_qgroup_order = [
        "BUSINESS_TYPE",
        "Respondent Profile",
        "Customer & Market",
        "Business & Strategy",
        "Pain Points & Needs",
        "Product & Process",
        "Product & Details",
        "Special Topic"
    ]

    already_handled = set()
    for group in preferred_qgroup_order:
        if group in grouped_questions_by_group:
            already_handled.add(group)
            for item in grouped_questions_by_group[group]:
                base_q, qty = item["question"], item["qty"]
                for i in range(1, qty + 1):
                    label = generate_unique_label(base_q, i, qty)
                    columns.append(label)
                    qgroup_row.append(group)
                    question_row.append(label)
                    pdf_rows.append([group, label, ""])

    for group in grouped_questions_by_group:
        if group not in already_handled and group != "N/A":
            for item in grouped_questions_by_group[group]:
                base_q, qty = item["question"], item["qty"]
                for i in range(1, qty + 1):
                    label = generate_unique_label(base_q, i, qty)
                    columns.append(label)
                    qgroup_row.append(group)
                    question_row.append(label)
                    pdf_rows.append([group, label, ""])

    for item in unmatched_questions:
        base_q, qty = item["question"], item["qty"]
        for i in range(1, qty + 1):
            label = generate_unique_label(base_q, i, qty)
            columns.append(label)
            qgroup_row.append("N/A")
            question_row.append(label)
            pdf_rows.append(["N/A", label, ""])

    # Cross product
    if is_cross and selected_products and selected_details:
        for prod in selected_products:
            for i in range(1, prod["qty"] + 1):
                for detail in selected_details:
                    label = generate_unique_label(f"{prod['name']}-{detail}", i, prod["qty"])
                    columns.append(label)
                    qgroup_row.append("Product & Details")
                    question_row.append(label)
                    pdf_rows.append(["Product & Details", label, ""])

    # ✅ สร้าง DataFrame สำหรับ Excel แนวนอน (หัว 2 แถว)
    header_df = pd.DataFrame([qgroup_row, question_row])
    empty = pd.DataFrame([[""] * len(columns) for _ in range(5)])
    final_df = pd.concat([header_df, empty], ignore_index=True)

    st.markdown("### 📓 ตัวอย่าง (Excel)")
    st.dataframe(final_df.head(5))

    # ✅ สร้าง Excel + ใส่ dropdown ให้คอลัมน์ยี่ห้อ/รุ่น/แบรนด์ ของกลุ่ม Product & Details
    excel_buffer = BytesIO()
    with pd.ExcelWriter(excel_buffer, engine="openpyxl") as writer:
        # เขียนตารางหลัก (หัว 2 แถว + Blank rows)
        final_df.to_excel(writer, sheet_name="Survey Template", index=False)
        ws = writer.sheets["Survey Template"]
        wb = writer.book
        # =========================
        # 1) เตรียม "พจนานุกรมตัวเลือก" ในชีต Dict
        # =========================

        # ---- 1.1: ลิสต์ของ product แยก grey mortar rmc ta tg
        GREY_PRODUCTS = [
            "01.Tiger","02.Rhino","03.Super","04.Tiger Plastering","06.Precast","11.Elephant Hybrid","20.Durable",
            "Corner Bead Mortar","FSM","GPM","LMM","LPM","MAM","Mortar Easy",
            "TPI-TPI Loft M103  ปูนฉาบขัดมันสำเร็จรูป สูตรผง ผสมน้ำใช้ได้ทันที",
            "TPI-TPI Loft Ready to used NP103 ผลิตภัณฑ์ฉาบขัดมันสำเร็จรูป สูตรพร้อมใช้",
            "TPI-คอนกรีตแห้งทีพีไอ (Dry Crete)",
            "TPI-ซีเมนต์แห้งเร็วพิเศษ ทีพีไอ (M680) (Water Plug Cement)",
            "TPI-ทีพีไอ ออยล์ เวล ซีเมนต์",
            "TPI-ปูนซีเมนต์ไฮดรอลิก ชนิดใช้งานทั่วไป ตราทีพีไอ 299",
            "TPI-ปูนซีเมนต์ปอร์ตแลนด์ประเภท 1 ตราทีพีไอ (สีแดง)",
            "TPI-ปูนซีเมนต์ปอร์ตแลนด์ประเภท 3 ตราทีพีไอ (สีดำ)",
            "TPI-ปูนซีเมนต์ปอร์ตแลนด์ประเภท 5 ตราทีพีไอ (สีฟ้า)",
            "TPI-ปูนซีเมนต์ผสม ตราทีพีไอ (สีเขียว)",
            "TPI-ปูนทีพีไอ เขียวซูเปอร์",
            "TPI-ปูนทีพีไอ M197/M199",
            "TPI-ปูนทีพีไอ แดงซูเปอร์",
            "TPI-ปูนสำเร็จรูปสำหรับงานทนกรด (M250)",
            "TPI-ปูนสำเร็จรูปสำหรับบล็อคมวลเบา (M220B) ชนิดไม่อบไอน้ำ",
            "Well Cement",
            "บัวเขียว","บัวแดง โปรเวิร์ค","บัวแดง ไฮเทค","บัวแดง ไฮเทค เอ็กซ์ตร้า","บัวแดง งานเททั่วไป งานหล่อ",
            "บัวฉลาม","บัวซูเปอร์","บัวดำ","บัวพลัส","บัวฟ้า",
            "บัวมอร์ตาร์ ก่อทั่วไป","บัวมอร์ตาร์ ก่ออิฐมวลเบา","บัวมอร์ตาร์ ฉาบทั่วไป","บัวมอร์ตาร์ ฉาบอิฐมวลเบา",
            "อินทรี ไพร์เมอร์","อินทรี ลาเท็กซ์","อินทรีเพชร","อินทรีเพชร CPM","อินทรีเพชร Easy Flow","อินทรีเพชร Quick Cast",
            "อินทรีเพชร งานทางหลวง","อินทรีเพชรพลัส","อินทรีแดง","อินทรีซูเปอร์",
            "อินทรีดำ High Early Strength","อินทรีดำ​","อินทรีดำงานหล่อ","อินทรีทอง","อินทรีปูนเขียว","อินทรีพ่น"
            
        ]
        MORTAR_PRODUCTS = [
            'Corner Bead Mortar-ปูนจับเซี๊ยม',' FSM-เสือมอร์ตาร์เทปรับพื้น','GPM-เสือมอร์ตาร์ฉาบทั่วไป',
            'LMM-เสือมอร์ตาร์ก่อมวลเบา','LPM-เสือมอร์ตาร์ฉาบมวลเบา',
            'MAM-เสือมอร์ตาร์ก่อทั่วไป',
            'Mortar Easy-เสือมอร์ตาร์ก่อเท',
            'Dry concrete-เสือมอร์ตาร์คอนกรีตแห้ง 240 KSC',
            'TPI-TPI Loft – M103  ปูนฉาบขัดมันสำเร็จรูป สูตรผง ผสมน้ำใช้ได้ทันที',
            'TPI-ซีเมนต์แห้งเร็วพิเศษ ทีพีไอ (M680) (Water Plug Cement)',
            'TPI-ปูนเทปรับระดับชนิดไหลตัวดี Semi-Self  M410',
            'TPI-ปูนเทปรับระดับสำเร็จรูป ทีพีไอ (M400)',
            'TPI-ปูนเทปรับระดับสำเร็จรูป ทีพีไอ (M409)',
            'TPI-ปูนก่อบล็อคมวลเบา ทีพีไอ',
            'TPI-ปูนก่อสำเร็จรูป ทีพีไอ',
            'TPI-ปูนฉาบบล็อคมวลเบา ทีพีไอ (M210)',
            'TPI-ปูนฉาบผิวคอนกรีต ทีพีไอ (M100C)',
            'TPI-ปูนฉาบละเอียดสำเร็จรูป ทีพีไอ (M100)',
            'TPI-ปูนฉาบสำเร็จรูปทั่วไป ทีพีไอ (M200)',
            'TPI-ปูนทีพีไอ M197/M199',
            'TPI-ปูนสำเร็จรูปสำหรับงานทนกรด (M250)',
            'TPI-ปูนสำเร็จรูปสำหรับบล็อคมวลเบา (M220B) ชนิดไม่อบไอน้ำ',
            'TPI-คอนกรีตแห้ง 240 KSC Cylinder (M402)',
            'บัวมอร์ตาร์ ก่อทั่วไป',
            'บัวมอร์ตาร์ ก่ออิฐมวลเบา',
            'บัวมอร์ตาร์ ฉาบทั่วไป',
            'บัวมอร์ตาร์ ฉาบอิฐมวลเบา',
            'อินทรีมอร์ตาร์ ฉาบทั่วไป 11',
            'อินทรีมอร์ตาร์ ฉาบละเอียด 12',
            'อินทรีมอร์ตาร์ ฉาบมวลเบา 13',
            'อินทรีมอร์ตาร์ ก่อทั่วไป 21',
            'อินทรีมอร์ตาร์ ก่อมวลเบา 23',
            'อินทรีมอร์ตาร์ เทปรับระดับพื้น 31',
            'อินทรีมอร์ตาร์ 52 คอนกรีตแห้ง 240 KSC'

        ]

        SKIM_PRODUCTS = [
            "Mass Grey skim coat","Mass White skim coat",
            "บัวมอร์ตาร์ สกิมโค้ท ปูนฉาบบาง แต่งผิว สีขาว","บัวมอร์ตาร์ สกิมโค้ท ปูนฉาบบาง ตกแต่งผิว สีเทา",
            "ซูเปอร์ สกิมโค้ท ทีพีไอ ผิวแกร่ง M651 (SUPER SKIM COAT HARDENING)",
            "ลูกดิ่ง สกิมโค้ท (สีขาว)","ลูกดิ่ง สกิมโค้ท (สีเทาอ่อน)","จระเข้ สกิมโค้ท สมูท",
            "ลูกดิ่ง สกิมโค้ท (สีเทา)","ลูกดิ่ง ซุปเปอร์ สกิมโค้ท  (สีขาว)",
            "ทีโอเอ 110 สกิมโค้ท สมูท เนื้อสีขาว","ทีโอเอ 110 สกิมโค้ท สมูท เนื้อสีเทา",
            "ทีโอเอ สกิมโค้ท เนื้อสีขาว​","ทีโอเอ สกิมโค้ท เนื้อสีเทา","จระเข้ สกิมโค้ท สมูท เกเตอร์",
            "LANKO สกิมโค้ท 110 สีเทา","ปูนฉาบผิวบาง Skim Coat TPI (M650F)",
            "จระเข้ สกิมโค้ท แซนด์ เกเตอร์","LANKO สกิมโค้ท 110 สีขาว","จระเข้ สกิมโค้ท 102","จระเข้ สกิมโค้ท 102 เกเตอร์"
        ]

        RMC_PRODUCTS = ['รถโม่ CPAC 210','รถโม่ CPAC 240', 'รถโม่ CPAC 280', 'รถโม่ CPAC 300',
                        'รถโม่ CPAC 320','รถโม่ SCG 210','รถโม่ SCG 240','รถโม่ SCG 280','รถโม่ SCG 300',
                        'รถโม่ SCG 320','รถโม่ อินทรีย์ 210','รถโม่ อินทรีย์ 240','รถโม่ อินทรีย์ 280','รถโม่ อินทรีย์ 300',
                        'รถโม่ อินทรีย์ 320','รถโม่ TPI 210','รถโม่ TPI 240','รถโม่ TPI 280','รถโม่ TPI 300','รถโม่ TPI 320'
            ]

        TA_PRODUCTS = [
            "Tile Adhesive Blue","Tile Adhesive Gold","Tile Adhesive Green","Tile Adhesive Orange","Tile Adhesive Pink",
            "COTTO TA","บัวมอร์ตาร์ กาวซีเมนต์ สำหรับกระเบื้องขนาดใหญ่","กาวซีเมนต์ ทีโอเอ ซิลเวอร์ไทล์",
            "บัวมอร์ตาร์ กาวซีเมนต์ สำหรับกระเบื้องทั่วไป","TPI-กาวซีเมนต์ ทีพีไอ (M500)",
            "TPI-กาวซีเมนต์ชนิดแรงยึดเกาะสูง ทีพีไอ (M501)",
            "TPI-ปูนกาวติดกระเบื้องขนาดใหญ่ สำหรับปูกระเบื้องสระว่ายน้ำ (M503)",
            "TPI-กาวซีเมนต์ชนิดพิเศษ (M509)","กาวซีเมนต์ ทีโอเอ พรีเมียมไทล์","กาวซีเมนต์เดฟโก้ ทีทีบีพลัส",
            "กาวซีเมนต์เดฟโก้ ซุปเปอร์ทีทีบี","กาวซีเมนต์เดฟโก้ แกรนิโต้ พลัส","กาวซีเมนต์เดฟโก้ พูล",
            "กาวซีเมนต์ จระเข้ทอง","กาวซีเมนต์ ทีโอเอ โปรไทล์","กาวซีเมนต์ จระเข้สโตนเมท",
            "กาวซีเมนต์ จระเข้เขียว","กาวซีเมนต์ จระเข้ฟ้า","กาวซีเมนต์ขาว จระเข้แดง","กาวซีเมนต์ จระเข้เอ็กซ์เพรส",
            "กาวซีเมนต์ ทีโอเอ อีโคไทล์","กาวซีเมนต์ จระเข้เกรย์สโตนเมท","กาวซีเมนต์ ทีโอเอ ซุปเปอร์ไทล์",
            "กาวซีเมนต์ จระเข้เอ็กซ์ตรีม","กาวซีเมนต์ขาว จระเข้ทอง","กาวซีเมนต์ จระเข้เงิน","กาวซีเมนต์ขาว จระเข้เงิน",
            "กาวซีเมนต์ จระเข้แดง","กาวซีเมนต์ เวเบอร์ไทล์ เฟล็กซ์","กาวซีเมนต์เดฟโก้ เอซี-2",
            "กาวซีเมนต์ เวเบอร์ไทล์ เกรส","กาวซีเมนต์ เวเบอร์ไทล์ วิส","กาวปูกระเบื้องพร้อมใช้ จระเข้ ทูฟิกซ์",
            "กาวปูและยาแนวกระเบื้อง จระเข้ อีพ็อกซี่ พลัส","กาวซีเมนต์ เวเบอร์สโตน ฟิกซ์","กาวซีเมนต์ เวเบอร์ไทล์ ฟิกซ์",
            "กาวซีเมนต์ จระเข้เหลือง","กาวซีเมนต์ จระเข้ทอง (สำหรับงานซ่อมแซม)",
            "กาวซีเมนต์ เวเบอร์ไทล์ 2-อิน-1","กาวซีเมนต์ เวเบอร์ไทล์ เซ็ม","กาวซีเมนต์ เวเบอร์ไทล์ โนสเตน",
            "กาวซีเมนต์ ชาละวัน","กาวซีเมนต์ จระเข้ทอง (สำหรับโมเสกแก้ว กระเบื้องแก้ว)","กาวซีเมนต์เดฟโก้ อัลตร้าเฟล็กซ์"
        ]

        TG_PRODUCTS = [
            "Tile Grout","กาวยาแนวอินทรี การ์ด","กาวยาแนวอินทรี โก",
            "TPI-ปูนสำเร็จรูป Non-Shrink Grout ชนิดไม่หดตัว (M670)",
            "TPI-ปูนเทไม่หดตัวสำหรับงานทั่วไป General Purpose Non-Shrink Grout M671",
            "TPI-ปูนยาแนว (Cement-Base Tile Grout) M550","TPI-กาวยาแนว ทีพีไอ คลาสสิค M552",
            "TPI-กาวยาแนว ทีพีไอ ซูเปอร์ พลัส  M551","กาวยาแนว จระเข้ บล็อกแก้ว","กาวยาแนว จระเข้ แพลทินัม",
            "กาวยาแนว จระเข้ เทอร์โบ พลัส","กาวยาแนว จระเข้ พรีเมี่ยม พลัส ทอง","กาวยาแนว จระเข้ พรีเมี่ยม พลัส เงิน",
            "กาวยาแนว เวเบอร์คัลเลอร์ พาวเวอร์","กาวยาแนว ชาละวัน","กาวปูและยาแนวกระเบื้อง จระเข้ อีพ็อกซี่ พลัส"
        ]

        PAINT_PRODUCTS = ["TOA","Beger","Nippon paint","Jotun","JBP","Dulux","Krystal"]

        # ---- 1.4: รวมเป็น dict สำหรับชีต Dict ----
        DICT_DATA = {
            "GREY":   GREY_PRODUCTS,
            "MORTAR": MORTAR_PRODUCTS,
            "SKIM":   SKIM_PRODUCTS,
            "TA":     TA_PRODUCTS,
            "TG":     TG_PRODUCTS,
            "RMC":    RMC_PRODUCTS,
            "PAINT":  PAINT_PRODUCTS,
        }

        dict_ws = wb.create_sheet("Dict")

        # เขียนหัวคอลัมน์
        for ci, cat in enumerate(DICT_DATA.keys(), start=1):
            dict_ws.cell(row=1, column=ci, value=cat)

        # เขียนรายการ
        max_len = 0
        for ci, cat in enumerate(DICT_DATA.keys(), start=1):
            items = DICT_DATA[cat]
            for ri, name in enumerate(items, start=2):
                dict_ws.cell(row=ri, column=ci, value=name)
            max_len = max(max_len, len(items))
        dict_ws.sheet_state = "visible"

        # =========================
        # 2) ใส่ Data Validation (อ้างช่วงตรงจากชีต Dict)
        # =========================
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.datavalidation import DataValidation
        from openpyxl.workbook.defined_name import DefinedName

        # รองรับ openpyxl หลายเวอร์ชัน
        def delete_named_range(wb, name: str):
            dn = wb.defined_names
            if hasattr(dn, "delete"):
                try: dn.delete(name)
                except Exception: pass
            else:
                try: dn.pop(name, None)
                except Exception:
                    try: del dn[name]
                    except Exception: pass

        def add_named_range(wb, name: str, ref: str):
            obj = DefinedName(name=name, attr_text=ref)  # workbook-scope
            dn = wb.defined_names
            if hasattr(dn, "add"):
                dn.add(obj)
            elif hasattr(dn, "append"):
                dn.append(obj)
            else:
                dn[name] = obj

        # 1) สร้าง Named Range ต่อกลุ่ม (LIST_GREY, LIST_MORTAR, ...), map เป็นชื่อ
        range_name_map = {}
        for ci, cat in enumerate(DICT_DATA.keys(), start=1):
            items = DICT_DATA[cat]
            if not items:  # ข้ามกลุ่มที่ไม่มีรายการ
                continue
            col_letter = get_column_letter(ci)
            end_row = len(items) + 1  # ข้อมูลเริ่มที่แถว 2 → แถวสุดท้าย = 1 + len
            nm = f"LIST_{cat}"        # เช่น LIST_GREY
            delete_named_range(wb, nm)
            # ใส่ quote ชื่อชีต กันชื่อแปลก/มีเว้นวรรค
            ref = f"'{dict_ws.title}'!${col_letter}$2:${col_letter}${end_row}"
            add_named_range(wb, nm, ref)
            range_name_map[cat] = nm

        # 2) ติด DV "ต่อคอลัมน์" (ไม่ reuse ต่อกลุ่ม) + เปิด in-cell dropdown
        BRAND_KEYS = ("ยี่ห้อ", "ยี่ห้อ/รุ่น", "รุ่น", "แบรนด์")
        HEADER_ROW = 3
        DATA_START_ROW = HEADER_ROW + 1  # = 4
        DATA_END_ROW = 100               # ปรับตามต้องการ

        def category_of_product(label: str) -> str | None:
            s = str(label).lower()
            if any(k in s for k in ["ยาแนว", " tile grout", "-tg", "mortar-tg", " tg-"]): return "TG"
            if any(k in s for k in ["กาวซีเมนต์", "tile adhesive", "-ta", "mortar-ta", " ta-"]): return "TA"
            if "skim" in s or "สกิม" in s or "mortar-สกิมโค้ท" in s: return "SKIM"
            if "paint" in s or "สี-" in s or s.startswith("สี-"): return "PAINT"
            if any(k in s for k in ["rmc", "ready mix", "ready-mix", "คอนกรีตผสมเสร็จ"]): return "RMC"
            if any(k in s for k in ["mortar", "มอร์ตาร์", "mortar-lw", "-lw", "lightweight"]): return "MORTAR"
            if any(k in s for k in ["grey", "เกรย์", "ปูนผง", "cement"]): return "GREY"
            return None

        max_col = ws.max_column
        for col_idx in range(1, max_col + 1):
            header_text = ws.cell(row=HEADER_ROW, column=col_idx).value
            if not header_text:
                continue
            if not any(k in str(header_text) for k in BRAND_KEYS):
                continue

            group = category_of_product(header_text)
            if group is None or group not in range_name_map:
                continue

            col_letter = get_column_letter(col_idx)
            cell_range = f"{col_letter}{DATA_START_ROW}:{col_letter}{DATA_END_ROW}"

            # ✅ In-cell dropdown ติ้กไว้ + allow blank + ไม่เด้ง error
            formula = f"={range_name_map[group]}"  # เช่น =LIST_GREY
            dv = DataValidation(type="list", formula1=formula, allow_blank=True)
            dv.showDropDown = False
            dv.allow_blank = True
            dv.showErrorMessage = False

            ws.add_data_validation(dv)
            dv.add(cell_range)

    # เสร็จใน writer; เซฟไฟล์นอก with หรือแปลง buffer ต่อได้
    # writer.save()  # ไม่จำเป็นใน context manager

    st.download_button("🔽️ ดาวน์โหลด Excel", data=excel_buffer.getvalue(),
                       file_name="survey_template.xlsx",
                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")



    # ✅ Preview PDF (ตารางตัวอย่าง)
    st.markdown("### 🔍 ตัวอย่าง (PDF)")
    st.dataframe(pd.DataFrame(pdf_rows[:5], columns=["Group", "Question", "Answer"]))

    # ฟอนต์ไทย (เช็คไฟล์ก่อนเพื่อกันพังตอนรันบนเครื่องที่ไม่มีฟอนต์)
    font_path = os.path.join("font", "THSarabun.ttf")
    if os.path.exists(font_path):
        pdfmetrics.registerFont(TTFont("THSarabun", font_path))
        font_name = "THSarabun"
        font_size = 14
    else:
        font_name = "Helvetica"
        font_size = 10
        st.warning("⚠️ ไม่พบฟอนต์ THSarabun.ttf — จะใช้ Helvetica แทนใน PDF")

    table_data = [["Group", "Question", "Answer"]] + pdf_rows
    row_heights = [25] + [60] * len(pdf_rows)

    pdf_buffer = BytesIO()
    doc = SimpleDocTemplate(pdf_buffer, pagesize=landscape(A4))
    table = Table(table_data, colWidths=[120, 280, 320], rowHeights=row_heights, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("FONTNAME", (0, 0), (-1, -1), font_name),
        ("FONTSIZE", (0, 0), (-1, -1), font_size),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
        ("GRID", (0, 0), (-1, -1), 1, colors.black),
    ]))
    doc.build([table])

    st.download_button("🔽️ ดาวน์โหลด PDF", data=pdf_buffer.getvalue(),
                       file_name="survey_questions_structured.pdf",
                       mime="application/pdf")

    # ✅ Excel แนวตั้ง (แบบ PDF) + ลำดับ
    df_vertical = pd.DataFrame(pdf_rows, columns=["Group", "Question", "Answer"])
    df_vertical.index += 1  # ให้เริ่มจาก 1
    df_vertical.reset_index(inplace=True)
    df_vertical.rename(columns={"index": "No."}, inplace=True)

    excel_vertical_buffer = BytesIO()
    with pd.ExcelWriter(excel_vertical_buffer, engine="openpyxl") as writer:
        df_vertical.to_excel(writer, sheet_name="Survey Vertical", index=False)

    st.download_button(
        label="⬇️ ดาวน์โหลด Excel (แนวตั้ง + ลำดับ)",
        data=excel_vertical_buffer.getvalue(),
        file_name="survey_template_vertical.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    # ✅ Preview Excel แนวตั้งใน Streamlit
    st.markdown("### 📋 ตัวอย่าง (Excel แนวตั้ง)")
    st.dataframe(df_vertical.head(10))

   # ✅ Excel สำหรับ Google Sheets (หัว 1 แถว, สะอาด, import ได้ทันที)
    gs_buffer = BytesIO()

    # ใช้เฉพาะคอลัมน์ที่เลือกไว้แล้วใน 'columns'
    gs_df = pd.DataFrame(columns=columns)

    with pd.ExcelWriter(gs_buffer, engine="openpyxl") as writer:
        # Sheet 1: Responses (ให้กรอกจริงใน Google Sheets)
        gs_df.to_excel(writer, sheet_name="Responses", index=False)
        ws = writer.sheets["Responses"]
        ws.freeze_panes = "A2"  # freeze หัวตาราง

        # Sheet 2: DataDictionary (อธิบายคอลัมน์ไว้ เผื่อใช้ใน AppSheet/ภายหลัง)
        dict_df = pd.DataFrame({
            "column_name": columns,
            "q_group": qgroup_row,
            "question_text": question_row,
        })
        dict_df.to_excel(writer, sheet_name="DataDictionary", index=False)

    st.download_button(
        label="⬇️ ดาวน์โหลด Excel (จำเป็นสำหรับใช้ใน Google Sheets)",
        data=gs_buffer.getvalue(),
        file_name="survey_google_sheets.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )





